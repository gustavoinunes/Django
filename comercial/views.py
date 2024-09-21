from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.views import View
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pandas as pd
import io

@login_required()
def tab_preco_venda(request):
    if request.method == 'POST':
        cod_produto = request.POST.get('processo')
        if cod_produto != '':
            return redirect('exporta_excel',cod_produto=cod_produto)
        else:
            return redirect('tab_preco_venda')
    else:
        return render(request, 'tabela_preco_venda.html')


class exporta_excel(View):
    def get(self, request,cod_produto):
        # Lê o arquivo CSV usando o delimitador correto
        df = pd.read_csv('C:/Users/gustavo.nunes/Desktop/data.csv', sep=';')
        
        # Transpõe a coluna VARIAVEL
        df_transposto = df.pivot_table(
            index=['TABELA', 'ITEM', 'MODULACAO', 'DESCRICAO', 'CARACTERISITCA'], 
            columns='VARIAVEL', 
            values='PRECO'
        )

        # Reset index para transformar o índice multi-nível em colunas normais
        df_transposto = df_transposto.reset_index()

        # Cria um buffer em memória para o arquivo Excel
        buffer = io.BytesIO()

        # Escreve o DataFrame no buffer como um arquivo Excel
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df_transposto.to_excel(writer, index=False, sheet_name='Dados Transpostos')
        
        # Move o cursor do buffer para o início
        buffer.seek(0)

        # Carrega o arquivo Excel do buffer
        book = load_workbook(filename=buffer)
        sheet = book.active

        # Define o estilo para a primeira linha (por exemplo, cor de fundo)
        fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        # Aplica o estilo à primeira linha
        for cell in sheet[1]:
            cell.fill = fill

        # Salva o arquivo alterado no buffer
        new_buffer = io.BytesIO()
        book.save(new_buffer)
        new_buffer.seek(0)

        # Cria uma resposta HTTP com o arquivo Excel
        response = HttpResponse(new_buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="dados_transpostos.xlsx"'

        return response